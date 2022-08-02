from docx import Document

document = Document('./test.docx')

from openpyxl import load_workbook, Workbook

wb = Workbook()
sheet = wb.active
sheet['A1'].value = 9999
sheet.cell()



